import numpy as np, pandas as pd
import re

def load_from_excel(filename, sheets=None, mask_hidden=True, use_openpyxl=True,
                    skip_rows=0,
                    **params): 
    """
    Reads the XLS file skipping masked rows.
    
    :param sheets       Sheet names to read data from
    :param skip_rows    Number of first lines to skip
    :param mask_hidden  If True, omit the masked lines
    """

    if sheets is not None:
        if isinstance(sheets, str):
            sheets = [sheets]
        elif not all(isinstance(sn, str) for sn in sheets):
            raise TypeError("Sheet names expected, not indices")

    # Find all masked lines and put them into sep. array
    if mask_hidden:
        skip_mask_rows = skip_rows + 1
        done_loading = False

        while not done_loading:
            # TODO normalize the sheet list prior to making the mask
            if use_openpyxl:
                from openpyxl import load_workbook
                from openpyxl.utils.exceptions import InvalidFileException
                try:
                    # FIXME 'row_dimensions' requires not to use 'read_only=True'
                    wb = load_workbook(filename=filename)
                except InvalidFileException:
                    use_openpyxl = False
                    continue

                sheet_masks = {}
                for sheetname in wb.get_sheet_names():
                    ws = wb.get_sheet_by_name(sheetname)
                    mask = np.ones(shape=(ws.max_row-skip_mask_rows,), dtype=bool)
                    for irow, row_info in ws.row_dimensions.items():
                        i = irow - skip_mask_rows - 1  # OpenPyxl counts rows from 1
                        if row_info.hidden and i >= 0:
                            mask[i] = False
                    sheet_masks[sheetname] = mask

                done_loading = True
            else: # use xlrd
                import xlrd

                # xlrd doesn't like sheet names with spaces
                sp2us = str.maketrans(' ', '_')
                sheets = [str.translate(sn, sp2us) for sn in sheets]

                with load_workbook(filename=filename, read_only=True, formatting_info=True) as book:
                    sheet_masks = {}
                    for s in book.sheets():
                        mask = np.ones(shape=(s.nrows-skip_mask_rows,), dtype=bool)
                        # TODO `rowinfo` doesn't indicate whether the row is hidden or not
                        for irow, row_info in s.rowinfo_map.items():
                            if row_info.hidden and irow >= skip_mask_rows:
                                mask[irow-skip_mask_rows] = False
                        sheet_masks[s.name] = mask
                done_loading = True
    
    sheets_data = pd.read_excel(filename, sheet_name=sheets, skiprows=skip_rows, **params)
    if mask_hidden:
        result = {}
        for name in sheets_data:
            data = sheets_data[name]
            mask = sheet_masks[name]
            n, m = data.shape[0], mask.shape[0]
            if n > m:
                mask = np.pad(mask, n, 'constant', constant_values=(True,))
            result[name] = data[mask[:data.shape[0]]]
        return result
    else:
        return sheets_data

def _split_column_names(column_names):
    if isinstance(column_names, str):
        return [cn.strip() for cn in re.split(r'[\t,; ]+', column_names) if cn]
    else:
        return column_names
    
def get_columns(df, column_names):
    return df[_split_column_names(column_names)].transpose().values

def get_columns_typed(df, column_names, dtype=np.float):
    column_names = _split_column_names(column_names)
    source = get_columns(df, column_names)
    dest = np.empty_like(source, dtype=dtype)
    it = np.nditer(source, flags=['multi_index', 'refs_ok'])
    def conv(v):
        if isinstance(v, str):
            return float(str.encode(v, 'latin1', errors='ignore'))
        else:
            return v

    while not it.finished:
        try:
            dest[it.multi_index] = conv(it[0])
        except ValueError as e:
            col, row = it.multi_index
            raise ValueError("Error in (%r, %r): %s" % (column_names[col], row+2, e)) from e
        it.iternext()
    return dest

def symmetrized(a):
    return a + a.T - np.diag(a.diagonal())

def make_cov_mats(disps, covs):
    """
    Makes covariance matric from given dispersions and correlation coeffs.
    
    Example:
    Given [a,b,c] -- three variables with dispersions [da,db,dc] resp.
    Then the correlation coeffs should be given like that:
    
     . | a | b | c |
     --+---+---+---+
     a |    Rab Rac 
     b |        Rbc    --->  [Rab, Rac, Rbc]
     c |            
     
     In other words, these coeffs are taken row-wise from upper triangular correlation matrix
     from left to right.
    """
    def make_cov_mat(disps, covs):
        import itertools as it
        C = np.ones(shape=(n,n), dtype=disps.dtype)
        for k, (i, j) in enumerate(it.combinations(range(n), 2)):
            if j != i:
                C[j,i] = C[i,j] = covs[k] * disps[i] * disps[j]
            else:
                C[i,i] = disps[i]**2
        return symmetrized(C)
    
    n = len(disps)
    if len(covs) != (n**2 - n)//2:
        raise ValueError("Covariant values array size mismatch: should be %d =(n**2 - n)//2, with n=%d"
                        % ((n**2 - n)//2, n))
    if 0:
        covs = np.swapaxes(np.array(covs).T, 0, 1)
        disps = np.swapaxes(np.array(disps).T, 0, 1)
        b = np.broadcast(disps, covs)
        out = np.empty(b.shape)
        out.flat = [make_cov_mat(ds, cs) for (ds, cs) in b]
        return out

    #disps = np.array(disps)

    covs = np.array(covs).T
    disps = np.array(disps).T
    for cov, disp in zip(covs, disps):
        C = np.zeros(shape=(n,n))
        ii = np.triu_indices(n, 1)
        C[ii] = C[(ii[1], ii[0])] = cov
        C[np.diag_indices(n)] = 1.0
        yield C * np.multiply.outer(disp, disp)
